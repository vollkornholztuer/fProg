import openpyxl
import math

wb = openpyxl.Workbook()

dateiName = "Datei16.xlsx"

ws1 = wb.active
ws1.title = "MeineErsteTabelle"


# range(-3.1, 3.1, 0.1)

x = -3.2

for row in range(2, 65, 1):
    x += 0.1
    function = x * math.sin(x)
    if (x < 1e-11 and x > 1e-12):
        x = 0
    ws1.cell(column=1, row=row, value=x)
    ws1.cell(column=2, row=row, value=function)

wb.save(filename=dateiName)
